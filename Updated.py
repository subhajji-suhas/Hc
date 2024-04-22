import streamlit as st
import os
import openpyxl

# Function for handling temporary file creation and management
def handle_temp_file(uploaded_file, temp_dir="./temp"):
    temp_file_path = os.path.join(temp_dir, uploaded_file.name)
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    try:
        # Write uploaded file to temporary location
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return temp_file_path
    except Exception as e:
        st.error(f"Error handling temporary file: {e}")
        return None

# File handling and editing subpart
def file_editing():
    st.subheader("File Editing")

    # File upload with unique key
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"], key="file_uploader")

    if uploaded_file is not None:
        temp_file_path = handle_temp_file(uploaded_file)
        if temp_file_path is None:
            st.stop()
            return

        try:
            # Load workbook
            workbook = openpyxl.load_workbook(temp_file_path)

            # Define sheets to be deleted by name
            sheets_to_delete = [
                "Change Log",
                "Info",
                "Threat Policy (Endpoint)",
                "Threat Policy (Server)",
                "Exceptions"
            ]

            # Delete specified sheets by name
            for sheet_name in sheets_to_delete:
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]

            # Rename "Global Settings" to "General Settings" if present
            if "Global Settings" in workbook.sheetnames:
                global_settings_sheet = workbook["Global Settings"]
                global_settings_sheet.title = "General Settings"
                global_settings_sheet["A1"] = "General Settings"

            # Delete "Calculations" sheet if present
            if "Calculations" in workbook.sheetnames:
                del workbook["Calculations"]

            # Multi-select box to choose additional sheets
            st.sidebar.subheader("Choose License for Sheets")
            include_endpoints = st.sidebar.checkbox("Endpoint License Sheets")
            include_servers = st.sidebar.checkbox("Server License Sheets")
            include_complete = st.sidebar.checkbox("Complete Lincense Sheets")

            if include_complete:
                include_endpoints = True
                include_servers = True
		global_settings_sheet["A6"] = "Live Response - Endpoint"
		global_settings_sheet["A7"] = "Live Response - Server"

            if include_endpoints:
                # Add sheets with "Endpoint" in the name
                endpoint_sheets = [sheet_name for sheet_name in workbook.sheetnames if "Endpoint" in sheet_name]
                for sheet_name in endpoint_sheets:
                    if sheet_name not in workbook.sheetnames:
                        workbook.create_sheet(title=sheet_name)

            if include_servers:
                # Add sheets with "Server" in the name
                server_sheets = [sheet_name for sheet_name in workbook.sheetnames if "Server" in sheet_name]
                for sheet_name in server_sheets:
                    if sheet_name not in workbook.sheetnames:
                        workbook.create_sheet(title=sheet_name)

            # Exclude sheets with "Endpoint" in the name
            if not include_endpoints and not include_complete:
                endpoint_sheets = [sheet_name for sheet_name in workbook.sheetnames if "Endpoint" in sheet_name]
                for sheet_name in endpoint_sheets:
                    if sheet_name in workbook.sheetnames:
                        del workbook[sheet_name]

            # Exclude sheets with "Server" in the name
            if not include_servers and not include_complete:
                server_sheets = [sheet_name for sheet_name in workbook.sheetnames if "Server" in sheet_name]
                for sheet_name in server_sheets:
                    if sheet_name in workbook.sheetnames:
                        del workbook[sheet_name]

            # Save the edited workbook
            edited_file_path = f"./temp/edited_{uploaded_file.name}"
            workbook.save(edited_file_path)

            # Download button
            try:
                with open(edited_file_path, 'rb') as f:
                    file_data = f.read()
                st.download_button(label='Dowload Updated File',
                                   data=file_data,
                                   file_name=f"edited_{uploaded_file.name}",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key='file_download_button',
                                   help='Right-click and save as...')
                os.remove(edited_file_path)
            except Exception as e:
                st.error(f'An error occurred while downloading the file: {e}')

        except Exception as e:
            st.error(f"Error reading or editing Excel file: {e}")
            os.remove(temp_file_path)  # Clean up temporary file on error

# Start the file editing process
file_editing()
